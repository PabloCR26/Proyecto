import libro
from libro import Libro
from usuario import Usuario
from empleado import Empleado
from colorama import Fore
import openpyxl
from fpdf import FPDF

#LISTAS Y DICCIONARIO PARA LOS DATOS DE PRESTAMO POR USUARIO
lista_de_usuarios = []
lista_de_empleados = []
lista_de_libros = []
dict_libros = {}

#MENU PRINCIPAL
def menu_ingreso():
    while True:
        opcion = int(input(' \n\t----------MENU--------------\n' +
                           '\n\t1. Añadir nuevo libro'
                           '\n\t2. Ver lista de libros'
                           '\n\t3. Buscar libro '
                           '\n\t4. Inscribir usuario'
                           '\n\t5. Inscribir empleado'
                           '\n\t6. Eliminar libros'
                           '\n\t7. Reservar libro'
                           '\n\t8. Filtrar libros'
                           '\n\t9. Modificar atributos del libro'
                           '\n\t10. Prestamo de libro'
                           '\n\t11. Devolver libro'
                           '\n\t12. Generar informe'
                           '\n\t13. Salir del sistema'
                           '\nIngrese la opcion que desea:\n'))

        if opcion == 1:
            anadir_libro()
        elif opcion == 2:
            mostrar_lista_libros()
        elif opcion == 3:
            manejar_busqueda()
        elif opcion == 4:
            crear_usuario()
        elif opcion == 5:
            cargar_empleados()
        elif opcion == 6:
            eliminar_libro()
        elif opcion == 7:
            reservar()
        elif opcion == 8:
            filtrar_libros_genero()
        elif opcion == 9:
            modificar_atributos_libro()
        elif opcion == 10:
            prestamo_libro()
        elif opcion == 11:
            devolucion_libro()
        elif opcion == 12:
            generar_informes()
        elif opcion == 13:
            break
        else:
            print('¡Opción no valida!')



#CREA UN OBJETO NUEVO DE LA CLASE USUARIO Y LO MUESTRA
def crear_usuario():
    cedula = input("Ingrese su numero de cedula: ")
    nombre = input("Ingrese su nombre: ")
    apellido1 = input("Ingrese su primer apellido: ")
    apellido2 = input("Ingrese su segundo apellido: ")
    direccion = input("Ingrese su direccion: ")
    telefono = input("Ingrese su numero de telefono: ")

    objUsuario = Usuario(cedula, nombre, apellido1, apellido2, direccion, telefono)
    lista_de_usuarios.append(objUsuario)
    print("Usuario: " + nombre + " agregado correctamente.")



#CARGA LOS EMPLEADOS DEL ARCHIVO TXT EN MODO LECTURA
#USA UN CICLO FOR QUE ITERA A TRAVÉS DE CADA LINEA DEL ARCHIVO
def cargar_empleados():
    with open('empleados.txt', 'r') as f:
        for linea in f:
            datos = linea.split(',')
            obj_empleado = Empleado(datos[0], datos[2], datos[3], datos[4], datos[1], datos[5])
            lista_de_empleados.append(obj_empleado)
            print('El empleado ' + obj_empleado.nombre + ' fue agregado correctamente')



#CREA UN NUEVO OBJETO DE LA CALSE LIBRO CON ID UNICO
def anadir_libro():
    titulo = input("Ingrese el titulo del libro: ")
    autor = input("Ingrese el nombre del autor: ")
    year = input("Ingrese el año de creacion: ")
    genero = input("Ingrese el genero: ")
    descripcion = input("Ingrese una breve descripcion del libro: ")

    objLibro = Libro(titulo, autor, year, genero, descripcion)
    lista_de_libros.append(objLibro)

    print("El libro " + titulo + " con el ID " + str(objLibro.id) + " fue agregado correctamente. ")



#ELIMINA EL LIBRO, SI EL ID DADO ES IGUAL AL ID DEL LIBRO
def eliminar_libro():
    id = input("Ingrese el ID del libro: ")
    for libro in lista_de_libros:
        if (str(libro.id) == id):
            lista_de_libros.remove(libro)
            print('El libro ' + id + ' ha sido eliminado.')



#BUSCA EL LBRO DE ACUERDO A LA PROPIEDAD A BUSCAR DEL MISM0
# #UTILIZA GETATTR PARA ACCEDER A LOS ATRIBUTOS
#UTILIZA UN BUCLE FOR QUE RECORRE CADA OBJETO LIBRO EN LA LISTA
def buscar_libro(propiedad_buscar):
    valor = input(f"Ingrese el valor de la propiedad {propiedad_buscar} del libro: ")
    for libro in lista_de_libros:
        if (str(getattr(libro, propiedad_buscar)) == valor):
            print(" ID: " + str(libro.id) + " \n Titulo: " + libro.titulo + "\n Autor: " + libro.autor
                  + "\n Genero: " + libro.genero + "\n Descripcion: " + libro.descripcion + "\n Año: " + libro.year)



#GESTIONA LA BÚSQUEDA DE LIBROS MEDIANTE LA PROPIEDAD A BUSCAR
def manejar_busqueda():
    opcion = int(input('\t----------MENU--------------\n'
                       '\n\t1. Buscar libro por título'
                       '\n\t2. Buscar libro por ID'
                       '\n\t3. Buscar libro por Autor'
                       '\n\t4. Buscar libro por año'
                       '\n\t5. Buscar libro por Genero'
                       '\nIngrese la opcion que desea:\n'))
    if opcion == 1:
        buscar_libro("titulo")
    elif opcion == 2:
        buscar_libro("id")
    elif opcion == 3:
        buscar_libro("autor")
    elif opcion == 4:
        buscar_libro("year")
    elif opcion == 5:
        buscar_libro("genero")
    else:
        print("Ingrese una opcion valida!!")



#SE USA LA FUNCIÓN FILTER PARA FILTRAR LA LISTA POR EL GÉNERO ESPECIFICADO
#LA FUNCIÓN LAMBDA VERIFICA EL GÉNERO DEL LIBRO COINCIDE CON EL VALOR INGRESADO
#SE USA LA FUNCIÓN LIST() SE USA PARA CONVERTIR EL RESULTADO EN UNA LISTA
def filtrar_libros_genero():
    valor = input('Ingrese el género del libro a filtrar: ')
    libros_filtrados = list(filter(lambda libro: libro.genero == valor, lista_de_libros))
    print('Los libros que coinciden son: ')
    for libros in libros_filtrados:
        print(libros.titulo)



#SE USA UN BUCLE FOR QUE RECORRE CADA LIBRO EN LA LISTA DE LIBROS
#SE COMPRUEBA SI EL ID INGRESADO ES CORRECTO, Y SE PROCEDE CON LA MODIFICACIÓN
#LOS IF COMPRUEBAN SI SE INGRESARON NUEVOS VALORES ARA LOS ATRIBUTOS
#SI SE AÑADE UN VALOR DIFERENTE A VACÍO, SE ACTUALIZA EL ATRIBUTO
def modificar_atributos_libro():
    id = input("Ingrese el ID del libro: ")
    titulo = input("Ingrese el titulo del libro: ")
    autor = input("Ingrese el nombre del autor: ")
    year = input("Ingrese el año de creacion: ")
    genero = input("Ingrese el genero: ")
    descripcion = input("Ingrese una breve descripcion del libro: ")

    for libro in lista_de_libros:
        if str(libro.id) == id:
            if (titulo != ""):
                libro.titulo = titulo
            if (autor != ""):
                libro.autor = autor
            if (year != ""):
                libro.year = year
            if (genero != ""):
                libro.genero = genero
            if (descripcion != ""):
                libro.descripcion = descripcion
            print(" Los atributos han sido modificados!!! ")



#SE USA UN FOR QUE RECORRE CADA LIBRO EN LA LISTA DE LIBROS
#SE IMPRIMEN LOS DETALLES DE CADA LIBRO LÍNEA POR LÍNEA
def mostrar_lista_libros():
    for libro in lista_de_libros:
        print(" ID: " + str(libro.id) + " \n Titulo: " + libro.titulo + "\n Autor: " + libro.autor
              + "\n Genero: " + libro.genero + "\n Descripcion: " + libro.descripcion + "\n Año: " + libro.year)



#SE VERIFICA SI EL ID INGRESADO COINCIDE CON EL DEL LIBRO
#SE VERIFICA SI EL LIBRO ESTÁ DISPONIBLE PARA PRÉSTAMO
#SE LLAMA AL MÉTODO PRESTAR DEL OBJETO LIBRO PARA MARCARLO COMO PRESTADO Y ESTABLECER FECHAS
#EL if cedula in dict libros VERIFICA SI EL USUARIO YA TIENE LIBROS PRESTADOS EN dict_libros
#SI EL USUARIO TIENE LIBROS PRESTADOS SE AÑADE A SU LISTA DE LIBROS PRESTADOS EN EL DICCIONARIO
#SI NO TIENE LIBROS PRESTADOS, SE CREA UNA NUEVA ENTRADA EN EL dict_libros CON LA CÉDULA Y LA LISTA CON EL LIBRO
def prestamo_libro():
    id = input("Ingrese el ID del libro que desea solicitar: ")
    cedula = input("Ingrese la cédula del usuario: ")
    fecha_prestamo = input("Ingrese fecha de inicio de prestamo del libro: ")
    fecha_devolucion = input("Ingrese fecha de devolucion del libro: ")
    for libro in lista_de_libros:
        if str(libro.id) == id:
            if libro.disponible:
                libro.prestar(fecha_prestamo, fecha_devolucion)
                print(type(dict_libros))
                if (len(dict_libros) > 0):
                    datos = dict_libros.get(cedula)
                    print(datos)
                    libros_usuarios = datos["libros"]
                    if (libros_usuarios):
                        libros_usuarios.append(libro)
                        dict_libros.update({"libros": libros_prestados})
                else:
                    dict_libros[cedula] = {"libros": [libro]}
                print("Dic: ", dict_libros.items())
                print("Su libro ha sido prestado correctamente")
            else:
                print("Su libro no se encuentra disponible!!")



#SE PIDE EL ID DEL LIBRO A RESERVAR Y LAS FECHAS
#SE RECORRE LA LISTA DE LIBROS Y SE COMPRUEBA SI EL ID COINCIDE
#if not libro.reservado VERIFICA SI EL LIBRO NO ESTÁ RESERVADO
#libro.reservar() LLAMA AL METODO RESERVAR PARA MARCARLO COMO RESERVADO
#SI ES LIBRO NO ESTÁ DISPONIBLE SE IMPRIME EN LA CONSOLA
def reservar():
    id = input("Ingrese el ID del libro que desea reservar : ")
    fecha_inicio = input('Ingrese la fecha de inicio de la reserva: ')
    fecha_final = input('Ingrese la fecha de fin de la reserva: ')
    for libro in lista_de_libros:
        if str(libro.id) == id:
            if not libro.reservado:
                libro.reservar(fecha_inicio, fecha_final)
                print('Su libro se ha reservado correctamente. ')
            else:
                print("Su libro no se encuentra disponible!")



#SE USA UN FOR PARA RECORRER CADA LIBRO EN LA LISTA DE LIBROS
#SE VERIFICA SI EL ID COINCIDE
#SE VERIFICA SI EL LIBRO NO ESTÁ DISPONIBLE (ESTÁ PRESTADO)
#libro.devolver() LLAMA AL METODO devolver() DE LA CLASE LIBRO PARA MARCARLO COMO DISPONIBLE
def devolucion_libro():
    id = input("Ingrese el ID del libro que desea devolver : ")
    for libro in lista_de_libros:
        if str(libro.id) == id:
            if not libro.disponible:
                libro.devolver()
            else:
                print("Su libro no se encuentra disponible!!")



#opcion_informe() SOLICITA LA OPCION DE INFORME A GENERAR EN EL MENU
#SE LLAMAN A LAS FUNCIONES PROPIAS DE GENERACIÓN DE INFORMES
def generar_informes():
    opcion_informe = int(input("1. Libros disponibles en la biblioteca\n"
                               "2. Inventario total\n"
                               "3. Prestamos vigentes\n"
                               "4. Usuarios registrados\n"
                               "5. Pestamos de libros por Usuarios\n"
                               "Ingrese lo que desea generar:\n"))
    if opcion_informe == 1:
        libros_disponibles()
    elif opcion_informe == 2:
        inventario_total()
    elif opcion_informe == 3:
        libros_prestados()
    elif opcion_informe == 4:
        usuarios_registrados()
    elif opcion_informe == 5:
        libros_prestados_usuarios()
    else:
        print("Ingrese una opcion valida!!")



#CREA UNA LISTA VACIA DONDE SE ALMACENARAN LOS LIBROS DISPONIBLES
#SE VERIFICA SI EL LIBRO ESTÁ DISPONIBLE CON LOS BUCLES Y SE AGREGA A LA LISTA
#SE LLAMA A LA FUNCION formato_generar CON LA LISTA COMO ARGUMENTO
def libros_disponibles():
    libros_disponibles = []
    for libro in lista_de_libros:
        if libro.disponible:
            libros_disponibles.append(libro)
    formato_generar(libros_disponibles)



#SE ALMACENAN LOS LIBROS PRESTADOS EN UNA LISTA
#SE VERIFICA SI EL LIBRO NO ESTÁ DISPONIBLE (ESTÁ PRESTADO)
#SI EL LIBRO ESTÁ PRESTADO, SE AGREGA A LA LISTA
#SE LLAMA A formato_generar Y SE PASA LA LISTA COMO ARGUMENTO PARA GENERAR EL INFORME
def libros_prestados():
    libros_prestados = []
    for libro in lista_de_libros:
        if not libro.disponible:
            libros_prestados.append(libro)
    formato_generar(libros_prestados)



#SE GENERA EL INFORME EN BASE A LA LISTA TOTAL DE LIBROS
def inventario_total():
    formato_generar(lista_de_libros)



#SE GENERA EL INFORME EN BASE A LA LISTA DE USUARIOS
def usuarios_registrados():
    formato_generar(lista_de_usuarios)


#SOLICITA LA CEDULA DE QUIEN SE DESEA GENERAR EL INFORME
# SE VERIFICA SI EL dict_libros CONTIENE DATOS
#usuario_libros = dict_libros.get(cedula) +
# OBTIENE LOS DATOS DEL USUARIO DEL DICCIONARIO CON LA CEDULA COMO CLAVE
#SE VERIFICA SI SE ENCONTRARON DATOS PARA EL USUARIO
#SE OBTIENE LA LISTA DE LIBROS PRESTADOS POR EL USUARIO
#SE VERIFICA SI TIENE LIBROS PRESTADOS
#SE CREA UNA LISTA temp_datos PARA ALMACENAR LOS LIBROS PRESTADOS POR EL USUARIO
def libros_prestados_usuarios():
    cedula = input("Ingrese la cédula del usuario: \n")
    if (len(dict_libros) > 0):
        usuario_libros = dict_libros.get(cedula)
        if (usuario_libros):
            libros = usuario_libros["libros"]
            if (len(libros) > 0):
                temp_datos = []
                for libro in libros:
                    temp_datos.append(
                        [cedula, str(libro.id), libro.titulo, libro.autor, libro.descripcion, libro.inicio_prestamo,
                         libro.fin_prestamo])
                formato_generar(temp_datos)



#FUNCION PARA GENERAR EL FORMATO DESEADO AL INFORME
def formato_generar(datos):
    opcion = int(input('Seleccione e el formato del archivo para guardar:\n'
                       '1. Archivo de texto\n'
                       '2. Archivo PDF \n'
                       '3. Archivo de Excel \n'
                       'Ingrese su opcion:'))

    if opcion == 1:
        generar_txt(datos)
    elif opcion == 2:
        generar_pdf(datos)
    elif opcion == 3:
        generar_excel(datos)
    else:
        print("Ingrese una opcion valida!!")


#ABRE EL ARCHIVO CON EL NOMBRE DADO Y LA EXTENSION TXT EN MODO ESCRITURA
def generar_txt(datos_guardar):
    nombre_archivo = input('Ingrese el nombre del archivo:')
    with open(nombre_archivo + '.txt', 'w') as f:
        for linea in datos_guardar:
            f.write(linea.__str__())
        print(Fore.RED + 'El archivo se guardó correctamente' + Fore.RESET)



#SE CREA UNA INSTANCIA DE LA CLASE FPDF PARA TRABAJAR CON EL ARCHIVO
def generar_pdf(datos_guardar):
    nombre_archivo = input('Ingrese el nombre del archivo: ')
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Arial', size=12)
    for nombre in (datos_guardar):
        pdf.cell(0, 5, txt=nombre.__str__(), ln=True)
    pdf.output(nombre_archivo + '.pdf')
    print(Fore.RED + 'El archivo se guardó correctamente' + Fore.RESET)


#CREA UNA INSTANCIA DE LA CLASE WORKBOOK PARA TRABAJAR CON EXCEL
#for i, nombre in enumerate... RECORRE CADA CONJUNTO DE DATOS, MIENTRAS i RASTREA EL NUMERO DE FILA
def generar_excel(datos_guardar):
    nombre_archivo = input("Ingrese el nombre del archivo")
    libro = openpyxl.Workbook()
    hoja = libro.active

    for i, nombre in enumerate(datos_guardar, start=1):
        hoja.cell(row=i, column=1, value=nombre.__str__())
    libro.save(nombre_archivo + ".xlsx")
    print(Fore.RED + 'El archivo se guardó correctamente' + Fore.RESET)


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    menu_ingreso()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
